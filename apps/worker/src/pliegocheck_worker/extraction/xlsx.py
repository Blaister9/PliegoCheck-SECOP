"""Extractor XLSX deterministico con openpyxl."""

from openpyxl import load_workbook

from pliegocheck_worker.extraction.limits import ExtractionLimits
from pliegocheck_worker.extraction.models import (
    ControlledExtractionError,
    ExtractionResultData,
    ExtractionWarningData,
    SegmentData,
)
from pliegocheck_worker.extraction.security import inspect_zip_container
from pliegocheck_worker.extraction.text import ensure_character_budget, normalize_text

ROWS_PER_SEGMENT = 25


def _cell_to_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\t", " ").replace("\n", " ").strip()


def extract_xlsx(path: str, limits: ExtractionLimits) -> ExtractionResultData:
    inspect_zip_container(path, limits)
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
    except Exception as exc:
        raise ControlledExtractionError(
            "EXTRACTION_FAILED", "No fue posible abrir el XLSX."
        ) from exc

    try:
        sheet_names = list(workbook.sheetnames)
        if len(sheet_names) > limits.max_sheets:
            raise ControlledExtractionError(
                "EXTRACTION_LIMIT_EXCEEDED",
                "El libro excede el limite de hojas configurado.",
            )
        segments: list[SegmentData] = []
        warnings: list[ExtractionWarningData] = []
        contains_formula = False
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            buffer: list[str] = []
            block_start = 1
            processed_rows = 0
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row_index > limits.max_rows_per_sheet:
                    warnings.append(
                        ExtractionWarningData(
                            code="XLSX_ROW_LIMIT_REACHED",
                            message="La hoja excede el limite de filas configurado.",
                            location={"sheet_name": sheet_name, "row": row_index},
                        )
                    )
                    break
                values = [_cell_to_text(value) for value in row]
                contains_formula = contains_formula or any(
                    value.startswith("=") for value in values
                )
                row_text = "\t".join(values).strip()
                if row_text:
                    if not buffer:
                        block_start = row_index
                    buffer.append(row_text)
                    if len(buffer) >= ROWS_PER_SEGMENT:
                        segments.append(
                            _segment(sheet_name, block_start, row_index, buffer, len(segments) + 1)
                        )
                        buffer = []
                processed_rows = row_index
            if buffer:
                segments.append(
                    _segment(sheet_name, block_start, processed_rows, buffer, len(segments) + 1)
                )
        if contains_formula:
            warnings.append(
                ExtractionWarningData(
                    code="XLSX_FORMULAS_NOT_EVALUATED",
                    message="El libro contiene formulas; se conserva el texto de la formula.",
                )
            )
        character_count = ensure_character_budget(
            (segment.text for segment in segments),
            limits.max_characters,
        )
        status = "COMPLETED_WITH_WARNINGS" if warnings else "COMPLETED"
        return ExtractionResultData(
            status=status,
            detected_format="xlsx",
            segments=segments,
            warnings=warnings,
            sheet_count=len(sheet_names),
            character_count=character_count,
        )
    finally:
        workbook.close()


def _segment(
    sheet_name: str,
    row_start: int,
    row_end: int,
    rows: list[str],
    sequence: int,
) -> SegmentData:
    return SegmentData(
        sequence=sequence,
        segment_type="SHEET_ROW",
        text=normalize_text("\n".join(rows)),
        sheet_name=sheet_name,
        row_start=row_start,
        row_end=row_end,
        source_location={"sheet_name": sheet_name, "row_start": row_start, "row_end": row_end},
    )
